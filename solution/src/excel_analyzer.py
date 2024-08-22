from openpyxl import load_workbook
import cell_model as cell_model
import utilities as util
import converters as converters


class ExcelAnalyzer:
    def __init__(self, path, sheet):
        self.file_path = path
        self.sheet_name = sheet

    def process_workbook(self):
        workbook = load_workbook(filename=self.file_path, data_only=False)
        sheet = workbook[self.sheet_name]
        models = ExcelAnalyzer._create_cell_models(sheet)
        ExcelAnalyzer.set_dependencies(models)

        # Calculate the values for all cells that don't have a formula
        for cell in models.values():
            if cell.value is None:
                cell.value = ExcelAnalyzer.calculate_value(models, cell.coordinate)

        workbook.close()
        return models

    @staticmethod
    def _create_cell_models(sheet):
        cell_models = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    model = converters.create_cell_model(cell)
                    cell_models[cell.coordinate] = model
        return cell_models

    @staticmethod
    def set_dependencies(models):
        for cell in models.values():
            if cell.formula:
                referenced_cells = util.get_referenced_cells(cell.formula)
                for ref_coordinate in referenced_cells:
                    ref_cell_model = models.get(ref_coordinate)
                    if ref_cell_model:
                        cell.add_dependency(ref_cell_model)

    @staticmethod
    def calculate_value(models, cell_coordinate):
        cell = models.get(cell_coordinate)
        if not cell:
            raise ValueError(f"Cell {cell_coordinate} not found.")
        if cell.value is not None and cell.formula is None:
            return cell.value
        if cell.formula is not None:
            ExcelAnalyzer.calculate_dependencies(cell)
            cell.value = ExcelAnalyzer.evaluate_formula(cell)
        return round(cell.value, 2)

    @staticmethod
    def calculate_dependencies(model):
        for dependency in model.dependencies:
            if dependency.value is None:
                dependency.value = ExcelAnalyzer.calculate_value(dependency.coordinate)
        return model

    @staticmethod
    def evaluate_formula(model):
        formula = model.formula
        context = {dep.coordinate: dep.value for dep in model.dependencies}
        return eval(str(util.parse_formula(formula, context)))

    @staticmethod
    def update_dependencies(models, cell_coordinate):
        model = models.get(cell_coordinate)
        if model and model.formula:
            model.dependencies.clear()
            referenced_cells = util.get_referenced_cells(model.formula)
            for ref_coordinate in referenced_cells:
                ref_cell_model = models.get(ref_coordinate)
                if ref_cell_model:
                    model.add_dependency(ref_cell_model)

    @staticmethod
    def recalculate_values(models, cell_coordinate):
        model = models.get(cell_coordinate)
        if model:
            model.value = ExcelAnalyzer.calculate_value(models, cell_coordinate)
            for dependent in models.values():
                if cell_model in dependent.dependencies:
                    ExcelAnalyzer.recalculate_values(models, dependent.coordinate)

    @staticmethod
    def update_value(models, cell_coordinate, new_value):
        model = models.get(cell_coordinate)
        if model:
            model.update_value(new_value)

            # Update dependencies and recalculate values
            for m in models.values():
                ExcelAnalyzer.update_dependencies(models, m.coordinate)
                ExcelAnalyzer.recalculate_values(models, m.coordinate)
